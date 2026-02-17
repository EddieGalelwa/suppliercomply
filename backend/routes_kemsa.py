"""
KEMSA Export Routes for SupplierComply
Handles KEMSA-compliant export file generation
"""

import logging
from datetime import datetime, timedelta
from flask import Blueprint, render_template, request, jsonify, send_file
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import csv
import io

# Import from extensions and models (no circular import issue)
from extensions import db
from models import Product, Activity

logger = logging.getLogger(__name__)
kemsa_bp = Blueprint('kemsa', __name__, url_prefix='/kemsa')


@kemsa_bp.route('/')
@login_required
def index():
    """Render KEMSA export page."""
    if not current_user.can_access():
        return jsonify({'success': False, 'error': 'Subscription required'}), 403
    return render_template('kemsa.html')


@kemsa_bp.route('/upload', methods=['POST'])
@login_required
def upload():
    """Handle CSV file upload for KEMSA export."""
    if not current_user.can_access():
        return jsonify({'success': False, 'error': 'Subscription required'}), 403
    
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if not file.filename.endswith('.csv'):
            return jsonify({'success': False, 'error': 'Please upload a CSV file'}), 400
        
        # Read CSV file - handle both bytes and string
        file_content = file.stream.read()
        if isinstance(file_content, bytes):
            file_content = file_content.decode("UTF-8")
        stream = io.StringIO(file_content, newline=None)
        csv_reader = csv.DictReader(stream)
        
        # Get original headers
        original_headers = csv_reader.fieldnames or []
        
        # Normalize headers: lowercase, replace spaces with underscores, remove special chars
        header_mapping = {}
        normalized_headers = []
        
        for header in original_headers:
            normalized = header.lower().strip().replace(' ', '_').replace('-', '_')
            # Remove any non-alphanumeric characters except underscores
            normalized = ''.join(c for c in normalized if c.isalnum() or c == '_')
            header_mapping[normalized] = header
            normalized_headers.append(normalized)
        
        # Read rows with normalized headers
        rows = []
        for i, row in enumerate(csv_reader):
            if i >= 5:  # Only first 5 rows for preview
                break
            
            # Create normalized row
            normalized_row = {}
            for norm_key, orig_key in header_mapping.items():
                normalized_row[norm_key] = row.get(orig_key, '')
            
            rows.append(normalized_row)
        
        # Count total rows
        stream.seek(0)
        total_rows = sum(1 for _ in csv.DictReader(io.StringIO(stream.read()))) - 1  # -1 for header
        
        # Generate file ID
        import uuid
        file_id = str(uuid.uuid4())
        
        # Generate suggested mappings based on common patterns
        suggested_mappings = {}
        for norm_header in normalized_headers:
            if 'product' in norm_header or 'name' in norm_header:
                suggested_mappings['product_name'] = norm_header
            elif 'quantity' in norm_header or 'qty' in norm_header:
                suggested_mappings['quantity'] = norm_header
            elif 'batch' in norm_header or 'lot' in norm_header:
                suggested_mappings['batch_number'] = norm_header
            elif 'expiry' in norm_header or 'expiration' in norm_header or 'date' in norm_header:
                suggested_mappings['expiry_date'] = norm_header
            elif 'unit' in norm_header or 'measure' in norm_header or 'uom' in norm_header:
                suggested_mappings['unit_of_measure'] = norm_header
            elif 'code' in norm_header or 'sku' in norm_header:
                suggested_mappings['product_code'] = norm_header
            elif 'barcode' in norm_header or 'gs1' in norm_header:
                suggested_mappings['gs1_barcode'] = norm_header
        
        # PRODUCTION: Only paid users can download (trial users can upload/preview only)
        can_download = current_user.is_paid()
        
        return jsonify({
            'success': True,
            'file_id': file_id,
            'headers': normalized_headers,
            'preview_rows': rows,
            'suggested_mappings': suggested_mappings,
            'can_download': can_download,
            'total_rows': max(0, total_rows)
        }), 200
        
    except Exception as e:
        logger.error(f"Upload error: {str(e)}")
        return jsonify({'success': False, 'error': 'Failed to process file'}), 500


@kemsa_bp.route('/preview', methods=['GET', 'POST'])
@login_required
def preview():
    """Preview products for KEMSA export."""
    if not current_user.can_access():
        return jsonify({'success': False, 'error': 'Subscription required'}), 403
    
    try:
        # If POST, process mappings from uploaded data
        if request.method == 'POST':
            data = request.get_json()
            mappings = data.get('mappings', {})
            
            # Validate required fields
            if not mappings.get('product_name'):
                return jsonify({
                    'success': False,
                    'error': 'Product Name mapping is required'
                }), 400
            
            # Get user's existing products as sample preview
            products = Product.query.filter_by(user_id=current_user.id).order_by(
                Product.created_at.desc()
            ).limit(10).all()
            
            preview_data = []
            for p in products:
                preview_data.append({
                    'product_name': p.name,
                    'quantity': p.quantity or 0,
                    'batch_number': p.batch_number or '',
                    'expiry_date': p.expiry_date.isoformat() if p.expiry_date else '',
                    'gtin': p.gtin or ''
                })
            
            # Validation results
            validation_errors = []
            if not preview_data:
                validation_errors.append('No products found to preview')
            
            # PRODUCTION: Only paid users can download
            can_download = current_user.is_paid()
            
            return jsonify({
                'success': True,
                'preview': preview_data,
                'total_rows': len(preview_data),
                'validation': {
                    'errors': validation_errors,
                    'warnings': []
                },
                'can_download': can_download
            }), 200
        
        # GET request - return user's products
        else:
            products = Product.query.filter_by(user_id=current_user.id).order_by(
                Product.created_at.desc()
            ).limit(10).all()
            
            return jsonify({
                'success': True,
                'products': [{
                    'id': p.id,
                    'name': p.name,
                    'batch_number': p.batch_number,
                    'expiry_date': p.expiry_date.isoformat() if p.expiry_date else None,
                    'quantity': p.quantity,
                    'gtin': p.gtin,
                    'barcode_url': p.barcode_url
                } for p in products],
                'total_count': Product.query.filter_by(user_id=current_user.id).count()
            }), 200
        
    except Exception as e:
        logger.error(f"Preview error: {str(e)}")
        return jsonify({'success': False, 'error': 'Failed to load preview'}), 500


@kemsa_bp.route('/download', methods=['POST'])
@login_required
def download():
    """Download KEMSA-formatted Excel file."""
    if not current_user.can_access():
        return jsonify({'success': False, 'error': 'Subscription required'}), 403
    
    # PRODUCTION: Only paid users can download
    if not current_user.is_paid():
        return jsonify({
            'success': False,
            'error': 'Excel download is a paid feature',
            'upgrade_required': True
        }), 403
    
    try:
        # Get user's products
        products = Product.query.filter_by(user_id=current_user.id).all()
        
        if not products:
            return jsonify({'success': False, 'error': 'No products to export'}), 404
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "KEMSA Export"
        
        # KEMSA format headers
        headers = ['Product Name', 'Quantity', 'Batch Number', 'Expiry Date', 'GTIN/Barcode']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for product in products:
            ws.append([
                product.name,
                product.quantity or 0,
                product.batch_number or '',
                product.expiry_date.strftime('%Y-%m-%d') if product.expiry_date else '',
                product.gtin or ''
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Log activity
        activity = Activity(
            user_id=current_user.id,
            action='kemsa_download',
            details=f'Downloaded {len(products)} products'
        )
        db.session.add(activity)
        db.session.commit()
        
        filename = f"KEMSA_Export_{current_user.company_name or 'Supplier'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Download error: {str(e)}")
        return jsonify({'success': False, 'error': 'Failed to generate download'}), 500


@kemsa_bp.route('/export', methods=['POST'])
@login_required
def export():
    """Generate KEMSA-compliant Excel export."""
    if not current_user.can_access():
        return jsonify({'success': False, 'error': 'Subscription required'}), 403
    
    try:
        data = request.get_json()
        export_type = data.get('type', 'full')  # 'full' or 'expiring'
        
        # Get products based on export type
        if export_type == 'expiring':
            days = data.get('days', 30)
            products = current_user.get_expiring_products(days)
        else:
            products = Product.query.filter_by(user_id=current_user.id).all()
        
        if not products:
            return jsonify({'success': False, 'error': 'No products found for export'}), 404
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "KEMSA Export"
        
        # Headers
        headers = ['Product Name', 'Batch Number', 'Expiry Date', 'Quantity', 'GTIN', 'Generated Date']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for product in products:
            ws.append([
                product.name,
                product.batch_number or '',
                product.expiry_date.isoformat() if product.expiry_date else '',
                product.quantity or 0,
                product.gtin or '',
                product.created_at.strftime('%Y-%m-%d') if product.created_at else ''
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Log activity
        activity = Activity(
            user_id=current_user.id,
            action='kemsa_export',
            details=f'Exported {len(products)} products, type: {export_type}'
        )
        db.session.add(activity)
        db.session.commit()
        
        filename = f"KEMSA_Export_{current_user.company_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"KEMSA export error: {str(e)}")
        return jsonify({'success': False, 'error': 'Failed to generate export'}), 500